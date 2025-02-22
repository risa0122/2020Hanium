import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import operator

menu = ['모짜맵구마핫도그', '봄쥬르김밥', '엽기오뎅', '119날치알주먹밥', '바른김밥',
        '응급실오뎅', '응급실닭볶음탕', '얌샘라면', '모다기', '엽기떡볶이', '참숯통닭발','엽기닭볶음탕',
        '엽기반반', '라돈모다기', '엽기닭발', '김선생비빔면', '응급실떡볶이', '크림치즈호두김밥',
        '매콤짱아찌김밥', '따따블치즈핫도그', '매운까르보나라파스타떡볶이', '스패니쉬오징어먹물마리',
        '바른멸치국수', '고구마통모짜핫도그', '감자통모짜핫도그', '숯불갈비핫도그', '119주먹밥',
        '듬뿍야채쫄쫄면', '베이컨마늘마리', '모짜렐라스팸계란마리','비빔모다기','버터갈릭감자튀김',
        '치믈렛떡볶이','수비드통닭떡볶이','통큰오짱떡볶이','오떡','돈까스오떡',
        '마라오떡','파삼겹오떡','오떡반반','찰떡순대','쿵쿵튀김','정글오뎅','킹콩파닭떡볶이','킹콩떡볶이',
        '더블크러스트이베리코','베스트콰트로','블랙앵거스스테이크','블랙타이거슈림프','와규앤비스테카',
        '러블리피스','새우천왕','올댓미트','치즈블라썸스테이크','트러플머쉬룸',
        '갈릭페퍼스테이크','수퍼파파스','스파이시치킨랜치','아메리칸핫도그피자','치볼레치킨피자',
        '더블포테이토피자','트리플바베큐피자','몬스터고구마피자','몬스터포테이토피자','쉬림프깐풍피자',
        '씨푸드새게피자','치즈핵폭탄피자','깐쇼새우피자','더블갈릭바베큐피자','도이치바이트피자',
        '맥시칸바이트피자','트러플버팔로피자','꿈을피자','어깨피자','웃음꽃피자','전주불백피자',
        '팔도피자','까망피자','맵고달고피자','반반한피자','퐁듀갈릭쉬림프','퐁듀불금피자',
        '갈릭버터쉬림프','더블치즈티본스테이크','뿜뿜불고기','순삭포테이토','치즈판타지',
        '6초치킨','간지치킨','고추치킨','60계양념치킨','장스치킨','60계후라이드','뱀파이어치킨',
        '소이갈릭스','자메이카통다리구이','치즐링','황금올리브치킨','갈비레오','마라칸','맛초킹',
        '맵스터','뿌링클','치하오','커리퀸','교촌레드콤보','교촌살살치킨','교촌오리지날','교촌허니콤보','교촌후라이드',
        '갈비천왕','고추바사삭','굽네오리지널','볼케이노치킨','허니멜로','쇼킹핫양념치킨','스노윙치킨',
        '오리엔탈파닭','포테이토짜용치킨','핫블링','네네치킨후라이드','단짠윙봉','볼빨간맵닭','어메이찡치킨',
        '오곡후라이드','콘듀치킨','더화이트치킨슈프림','슈프림양념치킨','와락치킨','치즈슈프림양념치킨',
        '치폴레양념치킨','푸라닭고추마요','레드홀릭치킨','블랙알리오치킨','파불로치킨']




filename = '분식ggul_data.xlsx'
wb = load_workbook(filename)
ws = wb.active

count = {}
for r in ws.rows:
         if r[2].value in menu:
                 if r[0].value != r[2].value:
                        count[r[0].value,r[2].value] = count.get((r[0].value,r[2].value),0) + 1

print(sorted(count.items(), key=operator.itemgetter(0), reverse=True))
